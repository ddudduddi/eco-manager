import io
import json
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from database import init_db, get_ecos_by_platform, get_platforms

init_db()

st.set_page_config(page_title="체크시트 생성", page_icon="📊", layout="wide")
st.title("체크시트 생성")

platforms = get_platforms()

if not platforms:
    st.warning("등록된 플랫폼이 없습니다. 먼저 플랫폼을 등록하세요.")
    st.stop()

selected_platform = st.selectbox("플랫폼 선택", options=platforms)

ecos = get_ecos_by_platform(selected_platform)

st.write(f"**{selected_platform}** 관련 ECO: **{len(ecos)}건**")

if not ecos:
    st.info("해당 플랫폼에 등록된 ECO가 없습니다.")
    st.stop()

# ── 미리보기 ──
df = pd.DataFrame(ecos)
df["platforms"] = df["platforms"].apply(
    lambda x: ", ".join(json.loads(x)) if x else ""
)
preview_cols = ["eco_number", "change_category", "change_summary", "apply_condition"]
rename_map = {
    "eco_number": "ECO 번호", "change_category": "변경 구분",
    "change_summary": "변경 내용 요약", "apply_condition": "적용 시점",
}
st.dataframe(
    df[preview_cols].rename(columns=rename_map),
    use_container_width=True,
    hide_index=True,
)

st.divider()


# ── Excel 생성 ──
def generate_checksheet(platform: str, eco_list: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "ECO 체크시트"

    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="맑은 고딕", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 타이틀 행
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"ECO 체크시트 - {platform}"
    title_cell.font = Font(name="맑은 고딕", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:F2")
    date_cell = ws["A2"]
    date_cell.value = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    date_cell.font = Font(name="맑은 고딕", size=9, color="666666")
    date_cell.alignment = Alignment(horizontal="right")

    # 헤더
    headers = ["No.", "ECO 번호", "변경 구분", "변경 내용 요약", "적용 시점",
               "검사결과", "비고"]
    col_widths = [6, 16, 14, 45, 18, 14, 20]

    header_row = 4
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 데이터 행
    for row_idx, eco in enumerate(eco_list, start=1):
        data_row = header_row + row_idx
        values = [
            row_idx,
            eco["eco_number"],
            eco["change_category"],
            eco["change_summary"],
            eco["apply_condition"],
            "",  # 검사결과 (빈칸)
            "",  # 비고 (빈칸)
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx in (1, 2, 3, 5, 6) else left_align

    # 검사결과 드롭다운 (Pass/Fail/N/A)
    dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,N/A"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="입력 오류",
        error="Pass, Fail, N/A 중 선택하세요.",
    )
    dv.prompt = "검사 결과를 선택하세요"
    dv.promptTitle = "검사결과"
    dv.showInputMessage = True

    last_data_row = header_row + len(eco_list)
    result_col = get_column_letter(6)  # F열 = 검사결과
    dv.add(f"{result_col}{header_row + 1}:{result_col}{last_data_row}")
    ws.add_data_validation(dv)

    # 인쇄 설정
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


filename = f"체크시트_{selected_platform}_{datetime.now().strftime('%Y%m%d')}.xlsx"

excel_data = generate_checksheet(selected_platform, ecos)

st.download_button(
    label="체크시트 Excel 다운로드",
    data=excel_data,
    file_name=filename,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
    type="primary",
)
